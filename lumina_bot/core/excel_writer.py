"""Excel output writer for processed fiscal documents."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Callable, Iterable, Literal

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from lumina_bot.core.logger import get_logger
from lumina_bot.models.nota import NotaFiscal


ExcelOutputMode = Literal["single_sheet", "multi_sheet", "one_file_per_pdf"]


class ExcelWriter:
    """Write fiscal document rows to Excel workbooks."""

    MONEY_KEYWORDS = (
        "valor",
        "iss",
        "inss",
        "pis",
        "cofins",
        "csll",
        "irrf",
        "retencoes",
        "descontos",
        "base_calculo",
        "aliquota",
    )
    DATE_KEYWORDS = ("data", "competencia")

    def __init__(self, output_path: Path) -> None:
        self._output_path = output_path
        self._logger = get_logger(self.__class__.__name__)

    def append(self, notas: Iterable[NotaFiscal]) -> None:
        """Append rows to a single-sheet workbook without dropping existing rows."""
        self.write(notas, mode="single_sheet")

    def write(
        self,
        notas: Iterable[NotaFiscal],
        *,
        mode: ExcelOutputMode = "single_sheet",
    ) -> None:
        """Write notes using the selected Excel output mode."""
        rows = list(notas)

        if not rows:
            return

        if mode == "multi_sheet":
            self.write_multi_sheet(rows)
            return

        if mode == "one_file_per_pdf":
            self.write_one_file_per_pdf(rows)
            return

        self.write_single_sheet(rows)

    def write_single_sheet(self, notas: Iterable[NotaFiscal]) -> None:
        """Append all documents to a single sheet."""
        rows = [nota.to_flat_dict() for nota in notas]

        if not rows:
            return

        self._output_path.parent.mkdir(parents=True, exist_ok=True)
        new_frame = pd.DataFrame(rows)

        if self._output_path.is_file():
            existing_frame = pd.read_excel(self._output_path, sheet_name="notas")
            frame = pd.concat([existing_frame, new_frame], ignore_index=True)
        else:
            frame = new_frame

        with pd.ExcelWriter(self._output_path, engine="openpyxl") as writer:
            frame.to_excel(writer, index=False, sheet_name="notas")

        self._format_workbook()
        self._logger.info("Excel updated: %s | new rows=%s", self._output_path, len(rows))

    def write_multi_sheet(
        self,
        notas: Iterable[NotaFiscal],
        sheet_name_resolver: Callable[[NotaFiscal], str] | None = None,
    ) -> None:
        """Write all documents into one workbook with one sheet per document."""
        rows = list(notas)

        if not rows:
            return

        self._output_path.parent.mkdir(parents=True, exist_ok=True)
        resolver = sheet_name_resolver or self._default_sheet_name
        used_names: set[str] = set()

        with pd.ExcelWriter(self._output_path, engine="openpyxl") as writer:
            for nota in rows:
                sheet_name = self._unique_sheet_name(resolver(nota), used_names)
                pd.DataFrame([nota.to_flat_dict()]).to_excel(
                    writer,
                    index=False,
                    sheet_name=sheet_name,
                )

        self._format_workbook()
        self._logger.info(
            "Multi-sheet Excel updated: %s | sheets=%s",
            self._output_path,
            len(rows),
        )

    def write_one_file_per_pdf(self, notas: Iterable[NotaFiscal]) -> None:
        """Create one Excel file for each processed document."""
        rows = list(notas)

        if not rows:
            return

        output_dir = self._output_path.parent
        output_dir.mkdir(parents=True, exist_ok=True)

        for nota in rows:
            file_name = self._safe_file_name(self._default_sheet_name(nota)) + ".xlsx"
            file_path = output_dir / file_name

            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                pd.DataFrame([nota.to_flat_dict()]).to_excel(
                    writer,
                    index=False,
                    sheet_name="nota",
                )

            self._format_workbook(file_path)

        self._logger.info("One-file-per-PDF Excel export completed: %s", output_dir)

    def _format_workbook(self, path: Path | None = None) -> None:
        workbook_path = path or self._output_path

        if not workbook_path.is_file():
            return

        workbook = load_workbook(workbook_path)

        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions

            for cell in sheet[1]:
                cell.font = Font(bold=True)

            for column_cells in sheet.columns:
                first_cell = column_cells[0]
                column_name = str(first_cell.value or "")
                column_letter = get_column_letter(first_cell.column)
                max_length = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in column_cells
                )
                sheet.column_dimensions[column_letter].width = min(
                    max(max_length + 2, 12),
                    80,
                )

                if self._is_money_column(column_name):
                    for cell in column_cells[1:]:
                        cell.number_format = '#,##0.00'

                if self._is_date_column(column_name):
                    for cell in column_cells[1:]:
                        cell.number_format = "dd/mm/yyyy"

        workbook.save(workbook_path)

    def _default_sheet_name(self, nota: NotaFiscal) -> str:
        candidates = (
            nota.prestador.nome_fantasia,
            nota.prestador.razao_social,
            nota.tomador.razao_social,
            nota.numero,
            nota.arquivo,
            "Nota",
        )

        for candidate in candidates:
            if candidate:
                return str(candidate)

        return "Nota"

    def _unique_sheet_name(self, desired_name: str, used_names: set[str]) -> str:
        base = self._safe_sheet_name(desired_name)
        name = base
        suffix = 2

        while name in used_names:
            suffix_text = f"_{suffix}"
            name = f"{base[:31 - len(suffix_text)]}{suffix_text}"
            suffix += 1

        used_names.add(name)
        return name

    @staticmethod
    def _safe_sheet_name(value: str) -> str:
        cleaned = re.sub(r"[\[\]:*?/\\]", " ", value).strip()
        cleaned = re.sub(r"\s+", " ", cleaned)
        return (cleaned or "Nota")[:31]

    @staticmethod
    def _safe_file_name(value: str) -> str:
        cleaned = re.sub(r"[^A-Za-z0-9_. -]+", "_", value).strip(" ._")
        return cleaned or "nota"

    def _is_money_column(self, column_name: str) -> bool:
        normalized = column_name.lower()
        return any(keyword in normalized for keyword in self.MONEY_KEYWORDS)

    def _is_date_column(self, column_name: str) -> bool:
        normalized = column_name.lower()
        return any(keyword in normalized for keyword in self.DATE_KEYWORDS)
